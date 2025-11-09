# app.py (debug-friendly, robust saving)
from flask import Flask, request, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os, traceback

# optional S3
import boto3
from botocore.exceptions import BotoCoreError, ClientError

app = Flask(__name__, static_folder='.', static_url_path='')

# Config via env
EXCEL_FILE = os.environ.get("EXCEL_FILE", "submissions.xlsx")
# make absolute path in repo cwd
EXCEL_PATH = os.path.join(os.getcwd(), EXCEL_FILE)
MULTIPLIER = int(os.environ.get("MULTIPLIER", "100"))

S3_BUCKET = os.environ.get("S3_BUCKET")
AWS_REGION = os.environ.get("AWS_REGION", None)

def ensure_dir_for_file(path):
    d = os.path.dirname(path)
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)

def ensure_excel():
    """Create the Excel file and headers if missing."""
    try:
        ensure_dir_for_file(EXCEL_PATH)
        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            headers = ['TimestampUTC', 'Name', 'Phone', 'Email', 'Zipcode', 'Bank', 'AmountSent', 'ReopenedFlag']
            ws.append(headers)
            wb.save(EXCEL_PATH)
        return True, None
    except Exception as e:
        return False, traceback.format_exc()

def append_row_and_save(row):
    """Append a row and save the workbook. Returns (ok, error_trace)"""
    try:
        ensure_dir_for_file(EXCEL_PATH)
        if not os.path.exists(EXCEL_PATH):
            # create file with headers if missing
            ok, err = ensure_excel()
            if not ok:
                return False, err
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        ws.append(row)
        wb.save(EXCEL_PATH)
        return True, None
    except Exception as e:
        return False, traceback.format_exc()

def upload_to_s3(local_path, bucket, key=None):
    if not bucket:
        return {"ok": False, "error": "No S3_BUCKET configured"}
    if key is None:
        ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        basename = os.path.basename(local_path)
        key = f"submissions/{ts}_{basename}"
    try:
        session_kwargs = {}
        if AWS_REGION:
            session_kwargs['region_name'] = AWS_REGION
        s3 = boto3.client('s3', **session_kwargs)
        s3.upload_file(local_path, bucket, key)
        return {"ok": True, "key": key, "bucket": bucket}
    except (BotoCoreError, ClientError) as e:
        return {"ok": False, "error": str(e), "trace": traceback.format_exc()}

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/status', methods=['GET'])
def status():
    """Admin endpoint for quick health check (no secrets)."""
    file_exists = os.path.exists(EXCEL_PATH)
    info = {
        "ok": True,
        "excel_path": EXCEL_PATH,
        "file_exists": file_exists,
        "MULTIPLIER": MULTIPLIER,
        "S3_BUCKET_configured": bool(S3_BUCKET)
    }
    return jsonify(info)

@app.route('/submit', methods=['POST'])
def submit():
    try:
        name = (request.form.get('name') or '').strip()
        phone = (request.form.get('phone') or '').strip()
        email = (request.form.get('email') or '').strip()
        zipcode = (request.form.get('zipcode') or '').strip()
        bank = (request.form.get('bank') or '').strip()
        amount_raw = (request.form.get('amount') or '').strip()
        reopened = (request.form.get('reopened') or '0').strip()

        # Basic validation
        if not name or len(name) < 2:
            return jsonify(success=False, error='Name too short'), 400
        if not phone or len(''.join(ch for ch in phone if ch.isdigit())) < 6:
            return jsonify(success=False, error='Phone looks invalid'), 400
        if not email or '@' not in email or '.' not in email.split('@')[-1]:
            return jsonify(success=False, error='Email looks invalid'), 400
        if not zipcode or len(zipcode) < 3:
            return jsonify(success=False, error='Zipcode looks invalid'), 400
        if not bank or len(bank) < 2:
            return jsonify(success=False, error='Bank name required'), 400

        try:
            raw_amount = float(amount_raw)
        except Exception:
            return jsonify(success=False, error='Amount is invalid'), 400

        # Apply multiplication server-side when reopened == '1'
        if reopened == '1':
            amount_sent = raw_amount * MULTIPLIER
            multiplied_flag = True
        else:
            amount_sent = raw_amount
            multiplied_flag = False

        timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

        row = [timestamp, name, phone, email, zipcode, bank, amount_sent, reopened]
        ok, err = append_row_and_save(row)
        file_exists = os.path.exists(EXCEL_PATH)

        s3_result = None
        if ok and S3_BUCKET:
            s3_result = upload_to_s3(EXCEL_PATH, S3_BUCKET)
        elif not S3_BUCKET:
            s3_result = {"ok": False, "error": "S3_BUCKET not configured; skipping upload"}

        if not ok:
            # Could not save file
            return jsonify(success=False, error='Failed to save Excel', save_trace=err, file_path=EXCEL_PATH), 500

        # Return informative JSON so frontend can display exact stored amount
        return jsonify(success=True, amount_sent=amount_sent, reopened_received=reopened, multiplied=multiplied_flag, file_exists=file_exists, file_path=EXCEL_PATH, s3_upload=s3_result)
    except Exception as exc:
        return jsonify(success=False, error=str(exc), trace=traceback.format_exc()), 500


@app.route('/download', methods=['GET'])
def download_excel():
    """Allow manual download of the current submissions.xlsx file."""
    try:
        path = os.path.join(os.getcwd(), "submissions.xlsx")
        if not os.path.exists(path):
            return jsonify(error="File not found. Submit once to create it."), 404
        return send_from_directory(os.getcwd(), "submissions.xlsx", as_attachment=True)
    except Exception as e:
        return jsonify(error=str(e)), 500


if __name__ == '__main__':
    ok, err = ensure_excel()
    if not ok:
        print("Failed to create excel on startup:", err)
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
